
import json
from openpyxl import load_workbook, Workbook

def create_rhumble():
    headers = ["id", "type", "name", "short name", "student rating", "credits", "course link", "stroke::", "radius::", "r::", "rel::dir::has prerequisite of", "rel::undir::has corequisite of", "rel::parent::belongs to", "fill::", "is_node::", "description"]
    row_data ={}
    # load prereq data 
    with open("./data/aggregated.json", "r") as f:
        for course_id, course_info in json.load(f).items():
            # skip all deparments except MATH and PHYS
            if course_info["subj"] not in ["MATH", "PHYS"]:
                continue
            row_data[course_id] = (
                {
                    "id": course_id,
                    "name": course_info["name"],
                    "rel::dir::has prerequisite of": "; ".join(course_info["prereqs"]),
                    "description": course_info["description"],
                    "rel::parent::belongs to": course_info["subj"],
                    "descendants": course_info["descendants"],
                }
            )

    max_immediate_descendants = 0
    for course, course_info in row_data.items():
        if len(course_info.get("descendants", [])) > max_immediate_descendants:
            max_immediate_descendants = len(course_info.get("descendants", []))

    assert max_immediate_descendants > 0

    # add formatting data to courses
    for course in row_data.keys():
        row_data[course]["type"] = "Course"
        row_data[course]["short name"] = row_data[course]["name"]
        row_data[course]["student rating"] = 0
        row_data[course]["credits"] = 4
        row_data[course]["course link"] = ""
        row_data[course]["stroke::"] = "#000000"
        #row_data[course]["radius::"] = 20
        row_data[course]["r::"] = 25 + round((25/max_immediate_descendants) * len(row_data[course].get("descendants", [])))
        row_data[course]["rel::undir::has corequisite of"] = ""
        row_data[course]["fill::"] = "#FFFFFF"
        row_data[course]["is_node::"] = "TRUE"

    
    # add departments
    departments = set()
    for course in row_data.keys():
        departments.add(row_data[course]["rel::parent::belongs to"])

    for department in departments:
        row_data[department] = (
            {
                "id": department,
                "type": "Department",
                "name": department,
                "short name": department,
                "rel::dir::has prerequisite of": "",
            }
        )

    # add root node
    row_data["root"] = (
        {
            "id": "root",
            "type": "root",
            "name": "RPI",
            "short name": "RPI",
            "rel::dir::has prerequisite of": "",
        }
    )

    

    wb = Workbook()
    wb.create_sheet("rhumble")
    ws = wb["rhumble"]

    # add headers 
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1, value=header)

    # add data
    for i, course_info in enumerate(row_data.items()):
        course_id, data = course_info
        for j, key in enumerate(headers):
            ws.cell(row=i+2, column=j+1, value=data.get(key, ""))

    # save to file
    wb.save("./rhumble.xlsx")

if __name__ == "__main__":
    create_rhumble()